from django.shortcuts import render, redirect, HttpResponse
from django.http import StreamingHttpResponse
from codetime.Extens import PageList
from codetime.models import Person, Product
from .forms import LoginForm, ProductForm
from . import models
from django.db.models import Sum
from openpyxl import load_workbook
from decimal import Decimal
import os

def jtbq(request):
    return render(request, 'codetime/jtbq.html')
def code(request):
    return render(request, 'codetime/code.html')
def index(request):
    return render(request, 'codetime/index.html')
# Create your views here.
def auth(func):
    def inner(reqeust, *args, **kwargs):
        v = reqeust.COOKIES.get('username111')
        if not v:
            return redirect('/codetime/login')
        return func(reqeust, *args, **kwargs)

    return inner


# 用户登录函数
def loginForm(request):
    if request.method == 'POST':
        name = request.POST.get('name', 'NAME')
        person = models.Person.objects.filter(name=name).first()
        login_form = LoginForm(request.POST)
        if person:
            password = request.POST.get('password', 'PASSWORD')
            if password == person.password:
                res = redirect('/codetime/product/0')
                res.set_cookie('username111', person.name)
                res.set_cookie('userlimitid', person.userlimitid)
                res.set_cookie('userId', person.id)
                return res
            return render(request, 'codetime/login.html')
        else:
            print(login_form.errors)
            return render(request, 'codetime/login.html', {"person": login_form.cleaned_data})
    return render(request, 'codetime/login.html')

@auth
def product(request, item_id):
    if str(item_id) == '0':
        return render(request, 'codetime/product.html')
    elif int(item_id) < 0:
        product = models.Product.objects.filter(id=abs(int(item_id))).first()
        product.id = '0'
        product.text3 = ''
        # product_form = ProductForm(instance=product)
        return render(request, 'codetime/product.html', {"product": product})
    else:
        product = models.Product.objects.filter(id=item_id).first()
    # product_form = ProductForm(instance=product)
        return render(request, 'codetime/product.html', {"product": product})

@auth
def againProduct(request, item_id):
    product = models.Product.objects.filter(id=item_id).first()
    product.id = '0'
    product.text3 = ''
    # product_form = ProductForm(instance=product)
    return render(request, 'codetime/product.html', {"product": product})

@auth
def productForm(request):
    if request.method == 'POST':
        product_form = ProductForm(request.POST)
        if product_form.is_valid():
            product_id = request.POST.get('product_id')
            unit = float(product_form.cleaned_data["text4"])
            number = int(product_form.cleaned_data["text8"])
            userId = request.COOKIES.get('userId')
            if product_id == "0":
                temp = product_form.save(commit=False)
                temp.amount = unit * number
                temp.userId = userId
                temp.save()
            else:
                model = models.Product.objects.filter(id=product_id).first()
                temp = product_form.save(commit=False)
                temp.amount = unit * number
                temp.id = product_id
                temp.text3 = model.text3
                temp.userId = userId
                temp.save()
            return render(request, 'codetime/product.html')
        else:
            print(product_form.errors)
            return render(request, 'codetime/product.html', {"product": product_form.cleaned_data})


@auth
def editProduct(request):
    limitId = request.COOKIES.get('userlimitid')
    if limitId == '0':
        return HttpResponse('OK')
    else:
        return  HttpResponse('False')


@auth
def changeShipments(request):
    limitId = request.COOKIES.get('userlimitid')
    if request.method == 'POST' and limitId == '0':
        item_id = int(request.POST.get('item_id'))
        shipments = request.POST.get('item_shipments')
        if shipments == 'True':
            models.Product.objects.filter(id=item_id).update(shipments=False)
        else:
            models.Product.objects.filter(id=item_id).update(shipments=True)
    return HttpResponse('OK')

#是否发货
@auth
def changeStatus(request):
    limitId = request.COOKIES.get('userlimitid')
    if request.method == 'POST' and limitId == '0':
        item_id = int(request.POST.get('item_id'))
        status = request.POST.get('item_status')
        if status == 'True':
            models.Product.objects.filter(id=item_id).update(status=False)
        else:
            models.Product.objects.filter(id=item_id).update(status=True)
    return HttpResponse('OK')


#是否结款
@auth
def changeKnot(request):
    limitId = request.COOKIES.get('userlimitid')
    if request.method == 'POST' and limitId == '0':
        item_id = int(request.POST.get('item_id'))
        knot = request.POST.get('item_knot')
        if knot == 'True':
            models.Product.objects.filter(id=item_id).update(knot=False)
        else:
            models.Product.objects.filter(id=item_id).update(knot=True)
    return HttpResponse('OK')


@auth
def orderSearch(request):
    condition = {}
    text1 = request.GET.get('text1', "")
    text2 = request.GET.get('text2', "")
    text23 = request.GET.get('text23', "")
    shipments = request.GET.get('shipments', "全部")
    status = request.GET.get('status', "全部")
    knot = request.GET.get('knot', "全部")
    # if len(text1) != 0:
    #     condition["text1"] = text1
    # if len(text2) != 0:
    #     condition["text2"] = text2
    if len(text23) != 0:
        condition["text24"] = text23
    if shipments != "全部":
        condition["shipments"] = int(shipments)
    if status != "全部":
        condition["status"] = int(status)
    if knot != "全部":
        condition["knot"] = int(knot)
    limitId = request.COOKIES.get('userlimitid')
    userId = request.COOKIES.get('userId')
    dataSource = models.Product.objects.all()
    if limitId != '0':
        dataSource = dataSource.filter(userId=userId)
    if len(dataSource) == 0:
        return render(request, 'codetime/orderSearch.html',{"nonReceiveAmount": 0, "receiveAmount": 0,"fullAmount": 0})
    dataSource = models.Product.objects.filter(**condition).order_by('-text3')
    if len(text1) != 0:
        dataSource = dataSource.filter(text1__icontains=text1)
    if len(text2) != 0:
        dataSource = dataSource.filter(text5__icontains=text2)
    if limitId != '0':
        dataSource = dataSource.filter(userId=userId)
    if len(dataSource) == 0:
        return render(request, 'codetime/orderSearch.html',{"nonReceiveAmount": 0, "receiveAmount": 0,"fullAmount": 0})
    current_page = request.GET.get('p', 1)
    current_page = int(current_page)
    val = 10
    page_obj = PageList.Page(current_page, len(dataSource), val)
    data = dataSource[page_obj.start:page_obj.end]

    page_str = page_obj.page_str("&text1="+text1+"&text2"+text2+"%text23&"+text23+"&shipments"+status+"&knot"+knot)
    nonReceiveAmount = 0
    receiveAmount = 0

    nonReceiveDic = dataSource.filter(knot=0).aggregate(amount=Sum('amount'))
    if nonReceiveDic["amount"]:
        nonReceiveAmount = nonReceiveDic["amount"].quantize(Decimal('0.00'))

    receiveDic = dataSource.filter(knot=1).aggregate(amount=Sum('amount'))
    if receiveDic["amount"]:
        receiveAmount = receiveDic["amount"].quantize(Decimal('0.00'))

    fullAmount = nonReceiveAmount + receiveAmount
    backDic = {}
    backDic["dataSource"] = data
    backDic["page_str"] = page_str
    if len(text1) != 0:
        condition["text1"] = text1
    if len(text2) != 0:
        condition["text2"] = text2
    backDic["condition"] = condition
    backDic["nonReceiveAmount"] = nonReceiveAmount
    backDic["receiveAmount"] = receiveAmount
    backDic["fullAmount"] = fullAmount
    # {"dataSource": data, 'page_str': page_str, "condition": condition, "nonReceiveAmount": 1000, "receiveAmount": 1000,
    #  "fullAmount": 2000}
    return render(request, 'codetime/orderSearch.html',backDic)


def writeOrderToExcel(request):
    item_id = int(request.GET.get('item_id'))
    product_obj = models.Product.objects.filter(id=item_id).first()

    # 默认可读写，若有需要可以指定write_only和read_only为True
    wb = load_workbook('productModel.xlsx')
    sheet = wb.active
    sheet.merge_cells('B3:C3')
    # 客户
    sheet['B3'] = product_obj.text1

    # 业务
    sheet['B4'] = product_obj.text2
    # 印刷版本号
    sheet['G3'] = product_obj.text3
    # sheet['G4'] = product_obj.amount
    # 产品名称
    sheet['B7'] = product_obj.text5
    sheet['B8'] = product_obj.text6
    sheet['B9'] = product_obj.text7
    sheet['B10'] = product_obj.text8
    sheet['B11'] = product_obj.text9
    sheet['B12'] = product_obj.text10
    sheet['B13'] = product_obj.text25
    sheet['B14'] = product_obj.text26
    sheet['B15'] = product_obj.text27
    sheet['B16'] = product_obj.text28
    sheet['B17'] = product_obj.text29
    # 版面枚数
    sheet['B18'] = product_obj.text30
    # 喷码
    sheet['E7'] = product_obj.text11
    sheet['E8'] = product_obj.text12
    sheet['E9'] = product_obj.text31
    sheet['E10'] = product_obj.text13
    sheet['E11'] = product_obj.text14
    sheet['E12'] = product_obj.text15
    sheet['E13'] = product_obj.text16
    sheet['E14'] = product_obj.text17
    sheet['E15'] = product_obj.text18
    sheet['E16'] = product_obj.text19
    # 菲林编号
    # sheet[''] = product_obj.text21
    # sheet['B30'] = product_obj.text20
    # 刀版编号
    sheet['E17'] = product_obj.text22
    sheet['E18'] = product_obj.retention_samples
    sheet['E3'] = product_obj.text23
    sheet['E4'] = product_obj.text24
    the_file_name = "tmp.xlsx"
    exist_file = os.path.exists(the_file_name)
    if exist_file:
        os.remove(r"tmp.xlsx")
    wb.save(the_file_name)

    def file_iterator(file_name, chunk_size=512):
        with open(file_name, 'rb') as f:
            while True:
                c = f.read(chunk_size)
                if c:
                    yield c
                else:
                    break

    response = StreamingHttpResponse(file_iterator(the_file_name))
    response['Content-Type'] = 'application/vnd.ms-excel'
    response['Content-Disposition'] = 'attachment;filename="{0}"'.format(the_file_name)
    return response

    # timestr = datetime.datetime.now().strftime("%Y%m%d%H%M%S")

    # system, node, release, version, machine, processor = platform.uname()
    # if system == 'Darwin':
    # print()
    # wb.save(os.getcwd() + timestr + '.xlsx')
    # else:
    #     path = timestr + '.xlsx'
    # wb.save(r'D:\\' + path)
    # print(system)
    # wb.save(r'D:\example.xlsx')
    # return HttpResponse('OK')



